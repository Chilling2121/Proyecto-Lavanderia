from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from .decorators import group_required
from django.db.models import Sum, Count, Q, ProtectedError
from django.utils import timezone
from django.http import HttpResponse
from django.contrib import messages
from django.db import transaction
from django.utils.dateparse import parse_datetime
from django.core.serializers.json import DjangoJSONEncoder
import re
import json
from decimal import Decimal, InvalidOperation
from django.contrib.auth.models import User, Group
from .models import Orden, Pago, PrendaOrden, Cliente, Servicio, SeguimientoLavado, TurnoCaja, MovimientoCaja, Promocion


# Flujo secuencial de estados de lavado
ESTADOS_LAVADO = [
    'Recibida',
    'Clasificada',
    'En lavado',
    'En secado',
    'En planchado',
    'Empaquetada',
    'Lista para entrega',
    'Entregada',
]

def validar_cedula_ecuatoriana(cedula):
    if cedula == '9999999999':
        return True
    if not re.match(r'^\d{10}$', cedula):
        return False
    
    provincia = int(cedula[:2])
    if not (1 <= provincia <= 24 or provincia == 30):
        return False
    
    tercer_digito = int(cedula[2])
    if tercer_digito >= 6:
        return False
    
    coeficientes = [2, 1, 2, 1, 2, 1, 2, 1, 2]
    suma = 0
    for i in range(9):
        val = int(cedula[i]) * coeficientes[i]
        if val >= 10:
            val -= 9
        suma += val
    
    verificador = int(cedula[9])
    residuo = suma % 10
    val_esperado = 0 if residuo == 0 else 10 - residuo
    
    return verificador == val_esperado

def home_redirect(request):
    if request.user.is_authenticated:
        if request.user.groups.filter(name='Operador').exists() and not request.user.groups.filter(name__in=['Administrador', 'Cajero']).exists():
            return redirect('seguimiento_list')
        return redirect('dashboard')
    return redirect('login')

@login_required
@group_required('Administrador', 'Cajero')
def dashboard(request):
    local_now = timezone.localtime(timezone.now())
    start_of_today = local_now.replace(hour=0, minute=0, second=0, microsecond=0)
    end_of_today = local_now.replace(hour=23, minute=59, second=59, microsecond=999999)
    from datetime import timedelta
    
    pedidos_hoy = Orden.objects.filter(fecha_recepcion__range=(start_of_today, end_of_today)).count()
    pedidos_pendientes = Orden.objects.exclude(estado_actual='Entregada').count()
    listos_entrega = Orden.objects.filter(estado_actual='Lista para entrega').count()
    ingresos_hoy = Pago.objects.filter(fecha_pago__range=(start_of_today, end_of_today)).aggregate(Sum('monto'))['monto__sum'] or 0.00
    
    ordenes_recientes = Orden.objects.select_related('cliente').order_by('-fecha_recepcion')[:5]
    
    total_prendas = PrendaOrden.objects.count()
    servicios_stats = []
    
    if total_prendas > 0:
        stats = PrendaOrden.objects.values('servicio__nombre').annotate(count=Count('id')).order_by('-count')
        for s in stats:
            percentage = int((s['count'] / total_prendas) * 100)
            servicios_stats.append({
                'nombre': s['servicio__nombre'],
                'percentage': percentage
            })
    else:
        servicios_stats = [
            {'nombre': 'Lavado Normal', 'percentage': 45},
            {'nombre': 'Lavado Delicado', 'percentage': 25},
            {'nombre': 'Planchado', 'percentage': 15},
            {'nombre': 'Lavado en Seco', 'percentage': 10},
            {'nombre': 'Otros', 'percentage': 5},
        ]
        
    # Datos para gráfico de barras (Últimos 7 días)
    dias_semana = []
    ingresos_semana = []
    for i in range(6, -1, -1):
        dia = start_of_today - timedelta(days=i)
        dia_end = end_of_today - timedelta(days=i)
        dias_semana.append(dia.strftime('%d/%m'))
        ingreso_dia = Pago.objects.filter(fecha_pago__range=(dia, dia_end)).aggregate(Sum('monto'))['monto__sum'] or 0.00
        ingresos_semana.append(float(ingreso_dia))

    # Datos para gráfico circular (Estados de las Órdenes activas)
    estado_labels = []
    estado_data = []
    estados_agrupados = Orden.objects.exclude(estado_actual='Entregada').values('estado_actual').annotate(total=Count('id'))
    for item in estados_agrupados:
        estado_labels.append(item['estado_actual'])
        estado_data.append(item['total'])
        
    context = {
        'pedidos_hoy': pedidos_hoy,
        'pedidos_pendientes': pedidos_pendientes,
        'listos_entrega': listos_entrega,
        'ingresos_hoy': ingresos_hoy,
        'ordenes_recientes': ordenes_recientes,
        'servicios_stats': servicios_stats,
        'ingresos_chart_labels': json.dumps(dias_semana),
        'ingresos_chart_data': json.dumps(ingresos_semana),
        'estado_chart_labels': json.dumps(estado_labels),
        'estado_chart_data': json.dumps(estado_data),
    }
    
    return render(request, 'dashboard.html', context)


# ==========================================
#  MÓDULO DE CLIENTES
# ==========================================

@login_required
@group_required('Administrador', 'Cajero')
def clientes_list(request):
    clientes = Cliente.objects.all().order_by('-id')
    return render(request, 'clientes.html', {'clientes': clientes})

@login_required
@group_required('Administrador', 'Cajero')
def clientes_search(request):
    query = request.GET.get('search', '').strip()
    if query:
        clientes = Cliente.objects.filter(
            Q(nombre__icontains=query) | Q(apellido__icontains=query) | Q(cedula__icontains=query) | Q(telefono__icontains=query)
        ).order_by('-id')
    else:
        clientes = Cliente.objects.all().order_by('-id')
    return render(request, 'partials/clientes_table.html', {'clientes': clientes})

@login_required
@group_required('Administrador', 'Cajero')
def cliente_historial(request, cliente_id):
    try:
        cliente = Cliente.objects.get(id=cliente_id)
        ordenes = Orden.objects.filter(cliente=cliente).order_by('-fecha_recepcion').prefetch_related('prendas__servicio')
    except Cliente.DoesNotExist:
        cliente = None
        ordenes = []
    return render(request, 'partials/cliente_historial.html', {'cliente': cliente, 'ordenes': ordenes})

@login_required
@group_required('Administrador', 'Cajero')
def cliente_create(request):
    if request.method == 'POST':
        cedula = request.POST.get('cedula', '').strip()
        nombre = request.POST.get('nombre', '').strip()
        apellido = request.POST.get('apellido', '').strip()
        telefono = request.POST.get('telefono', '').strip()
        correo = request.POST.get('correo')
        direccion = request.POST.get('direccion')
        
        # Validaciones
        if not validar_cedula_ecuatoriana(cedula):
            messages.error(request, "Error al registrar cliente: La cédula ingresada no es válida.")
            return redirect('clientes_list')
            
        if cedula != '9999999999' and Cliente.objects.filter(cedula=cedula).exists():
            messages.error(request, "Error al registrar cliente: La cédula ya está registrada para otro cliente.")
            return redirect('clientes_list')
            
        if not nombre or not apellido:
            messages.error(request, "Error al registrar cliente: El nombre y el apellido son obligatorios.")
            return redirect('clientes_list')
        
        if not re.match(r'^\d{10}$', telefono):
            messages.error(request, "Error al registrar cliente: El número de celular debe tener exactamente 10 dígitos.")
            return redirect('clientes_list')
        
        Cliente.objects.create(
            cedula=cedula,
            nombre=nombre,
            apellido=apellido,
            telefono=telefono,
            correo=correo or None,
            direccion=direccion or None
        )
        messages.success(request, f"Cliente '{nombre} {apellido}' registrado con éxito.")
    return redirect('clientes_list')

@login_required
@group_required('Administrador', 'Cajero')
def cliente_edit(request, cliente_id):
    if request.method == 'POST':
        try:
            cliente = Cliente.objects.get(id=cliente_id)
            cedula = request.POST.get('cedula', '').strip()
            nombre = request.POST.get('nombre', '').strip()
            apellido = request.POST.get('apellido', '').strip()
            telefono = request.POST.get('telefono', '').strip()
            correo = request.POST.get('correo')
            direccion = request.POST.get('direccion')
            
            # Validaciones
            if not validar_cedula_ecuatoriana(cedula):
                messages.error(request, "Error al editar cliente: La cédula ingresada no es válida.")
                return redirect('clientes_list')
                
            if cedula != '9999999999' and Cliente.objects.filter(cedula=cedula).exclude(id=cliente.id).exists():
                messages.error(request, "Error al editar cliente: La cédula ya está registrada para otro cliente.")
                return redirect('clientes_list')
                
            if not nombre or not apellido:
                messages.error(request, "Error al editar cliente: El nombre y el apellido son obligatorios.")
                return redirect('clientes_list')
            
            if not re.match(r'^\d{10}$', telefono):
                messages.error(request, "Error al editar cliente: El número de celular debe tener exactamente 10 dígitos.")
                return redirect('clientes_list')
            
            cliente.cedula = cedula
            cliente.nombre = nombre
            cliente.apellido = apellido
            cliente.telefono = telefono
            cliente.correo = correo or None
            cliente.direccion = direccion or None
            cliente.save()
            messages.success(request, f"Datos del cliente '{cliente.nombre_completo}' actualizados con éxito.")
        except Cliente.DoesNotExist:
            messages.error(request, "Error al editar cliente: El cliente especificado no existe.")
    return redirect('clientes_list')

@login_required
@group_required('Administrador')
def cliente_delete(request, cliente_id):
    if request.method == 'DELETE':
        try:
            cliente = Cliente.objects.get(id=cliente_id)
            nombre = cliente.nombre
            cliente.delete()
            response = HttpResponse("")
            response['HX-Trigger'] = '{"clientDeleted": "' + nombre + '"}'
            return response
        except Cliente.DoesNotExist:
            return HttpResponse(status=404)
    return redirect('clientes_list')


# ==========================================
#  MÓDULO DE SERVICIOS
# ==========================================

@login_required
@group_required('Administrador', 'Cajero')
def servicios_list(request):
    servicios = Servicio.objects.all().order_by('nombre')
    return render(request, 'servicios.html', {'servicios': servicios})

@login_required
@group_required('Administrador', 'Cajero')
def servicios_search_api(request):
    query = request.GET.get('search', '').strip()
    if query:
        servicios = Servicio.objects.filter(
            Q(nombre__icontains=query) | Q(descripcion__icontains=query)
        ).order_by('nombre')
    else:
        servicios = Servicio.objects.all().order_by('nombre')
    return render(request, 'partials/servicios_table.html', {'servicios': servicios})

@login_required
@group_required('Administrador')
def servicio_create(request):
    if request.method == 'POST':
        nombre = request.POST.get('nombre')
        descripcion = request.POST.get('descripcion')
        tarifa = request.POST.get('tarifa')
        tipo_cobro = request.POST.get('tipo_cobro')
        
        Servicio.objects.create(
            nombre=nombre,
            descripcion=descripcion,
            tarifa=tarifa,
            tipo_cobro=tipo_cobro
        )
        messages.success(request, f"Servicio '{nombre}' creado con éxito.")
        return redirect('servicios_list')
    
    return render(request, 'servicio_form.html')

@login_required
@group_required('Administrador')
def servicio_edit(request, servicio_id):
    try:
        servicio = Servicio.objects.get(id=servicio_id)
    except Servicio.DoesNotExist:
        messages.error(request, "El servicio solicitado no existe.")
        return redirect('servicios_list')
    
    if request.method == 'POST':
        nombre = request.POST.get('nombre')
        descripcion = request.POST.get('descripcion')
        tarifa = request.POST.get('tarifa')
        tipo_cobro = request.POST.get('tipo_cobro')
        
        servicio.nombre = nombre
        servicio.descripcion = descripcion
        servicio.tarifa = tarifa
        servicio.tipo_cobro = tipo_cobro
        servicio.save()
        
        messages.success(request, f"Servicio '{nombre}' actualizado con éxito.")
        return redirect('servicios_list')
        
    return render(request, 'servicio_form.html', {'servicio': servicio})

@login_required
@group_required('Administrador')
def servicio_delete(request, servicio_id):
    if request.method == 'DELETE':
        try:
            servicio = Servicio.objects.get(id=servicio_id)
            nombre = servicio.nombre
            servicio.delete()
            response = HttpResponse("")
            response['HX-Trigger'] = '{"servicioDeleted": "' + nombre + '"}'
            return response
        except ProtectedError:
            return HttpResponse("Error: Este servicio está vinculado a órdenes existentes y no puede ser eliminado por seguridad. Puede modificar su tarifa en su lugar.", status=400)
        except Servicio.DoesNotExist:
            return HttpResponse(status=404)
    return redirect('servicios_list')


# ==========================================
#  MÓDULO DE ÓRDENES
# ==========================================

@login_required
@group_required('Administrador', 'Cajero')
def ordenes_list(request):
    estado = request.GET.get('estado', '').strip()
    ordenes = Orden.objects.select_related('cliente').order_by('-id')
    if estado:
        ordenes = ordenes.filter(estado_actual=estado)
    return render(request, 'ordenes.html', {
        'ordenes': ordenes,
        'estados': ESTADOS_LAVADO,
        'estado_preseleccionado': estado,
    })

@login_required
@group_required('Administrador', 'Cajero')
def ordenes_search(request):
    query = request.GET.get('search', '').strip()
    estado = request.GET.get('estado', '').strip()
    
    ordenes = Orden.objects.select_related('cliente').order_by('-id')
    
    if query:
        id_filter = Q()
        if query.isdigit():
            id_filter = Q(id=int(query))
        ordenes = ordenes.filter(id_filter | Q(cliente__nombre__icontains=query))
        
    if estado:
        ordenes = ordenes.filter(estado_actual=estado)
        
    return render(request, 'partials/ordenes_table.html', {'ordenes': ordenes})

@login_required
@group_required('Administrador', 'Cajero')
def orden_cliente_search_api(request):
    query = request.GET.get('search', '').strip()
    if query:
        clientes = Cliente.objects.filter(
            Q(nombre__icontains=query) | Q(apellido__icontains=query) | Q(cedula__icontains=query) | Q(telefono__icontains=query)
        ).order_by('-id')[:5]
    else:
        clientes = []
    return render(request, 'partials/orden_cliente_results.html', {'clientes': clientes})

@login_required
@group_required('Administrador', 'Cajero')
def orden_create(request):
    from .models import CatalogoPrenda, Promocion
    from django.utils import timezone
    today = timezone.localtime(timezone.now()).date()
    
    if request.method == 'POST':
        cliente_id = request.POST.get('cliente_id')
        fecha_entrega_estimada_str = request.POST.get('fecha_entrega_estimada')
        observaciones = request.POST.get('observaciones', '')
        promocion_id = request.POST.get('promocion_id')
        
        promocion = None
        if promocion_id:
            try:
                promocion = Promocion.objects.filter(id=promocion_id).first()
            except Exception:
                pass
        
        try:
            total = Decimal(request.POST.get('total', '0.00') or '0.00')
            descuento = Decimal(request.POST.get('descuento', '0.00') or '0.00')
            total_a_pagar = Decimal(request.POST.get('total_a_pagar', '0.00') or '0.00')
            anticipo = Decimal(request.POST.get('anticipo', '0.00') or '0.00')
        except (InvalidOperation, TypeError, ValueError):
            messages.error(request, "Error al procesar los montos de la orden.")
            return redirect('ordenes_list')
            
        metodo_pago = request.POST.get('metodo_pago', 'Efectivo')
        prendas_json = request.POST.get('prendas_json', '[]')
        
        # Validar turno de caja antes de hacer cualquier cambio
        turno_activo = TurnoCaja.objects.filter(estado='Abierta').first()
        if not turno_activo:
            messages.error(request, "No puedes registrar órdenes nuevas sin abrir un turno de caja primero.")
            return redirect('orden_create')
        
        if not cliente_id:
            messages.error(request, "Debe seleccionar un cliente para la orden de servicio.")
            return redirect('orden_create')
            
        try:
            prendas_data = json.loads(prendas_json)
        except json.JSONDecodeError:
            messages.error(request, "Error en el formato de los detalles de las prendas.")
            return redirect('orden_create')
            
        if not prendas_data:
            messages.error(request, "Debe agregar al menos una prenda a la orden.")
            return redirect('orden_create')
        
        fecha_entrega_estimada = None
        if fecha_entrega_estimada_str:
            fecha_entrega_estimada = parse_datetime(fecha_entrega_estimada_str)
            if fecha_entrega_estimada and timezone.is_naive(fecha_entrega_estimada):
                fecha_entrega_estimada = timezone.make_aware(fecha_entrega_estimada)
        
        try:
            with transaction.atomic():
                cliente = Cliente.objects.get(id=cliente_id)
                
                orden = Orden.objects.create(
                    cliente=cliente,
                    promocion=promocion,
                    fecha_entrega_estimada=fecha_entrega_estimada,
                    total=total,
                    descuento=descuento,
                    total_a_pagar=total_a_pagar,
                    anticipo=anticipo,
                    saldo_pendiente=total_a_pagar - anticipo,
                    estado_pago='Liquidado' if anticipo >= total_a_pagar else ('Parcial' if anticipo > 0 else 'Pendiente'),
                    estado_actual='Recibida',
                    observaciones=observaciones or None
                )
                
                SeguimientoLavado.objects.create(
                    orden=orden,
                    estado='Recibida',
                    observaciones='Ingreso inicial de la orden de servicio al sistema.'
                )
                
                for p in prendas_data:
                    servicio = Servicio.objects.get(id=p['servicio_id'])
                    tipo_prenda_nombre = p['tipo_prenda'].strip()
                    if tipo_prenda_nombre:
                        CatalogoPrenda.objects.get_or_create(nombre=tipo_prenda_nombre)
                    # Leer el precio unitario si se envía tarifa personalizada
                    precio_unitario = float(p['precio_unitario']) if p.get('precio_unitario') is not None else None
                    PrendaOrden.objects.create(
                        orden=orden,
                        tipo_prenda=tipo_prenda_nombre,
                        cantidad=int(p.get('cantidad', 1)),
                        peso=float(p['peso']) if p.get('peso') else None,
                        servicio=servicio,
                        precio_unitario=precio_unitario,
                        es_delicada=bool(p.get('es_delicada', False)),
                        observaciones=p.get('observaciones', '') or None
                    )
                
                if anticipo > 0:
                    Pago.objects.create(
                        orden=orden,
                        monto=anticipo,
                        metodo_pago=metodo_pago,
                        tipo_pago='Anticipo',
                        observaciones='Abono inicial al registrar la orden',
                        turno=turno_activo
                    )
                
                messages.success(request, f"Orden de servicio #{orden.id:06d} registrada con éxito.")
                return redirect('ordenes_list')
        except Cliente.DoesNotExist:
            messages.error(request, "El cliente seleccionado no existe.")
        except Exception as e:
            messages.error(request, f"Ocurrió un error inesperado al procesar la orden: {str(e)}")
            
        return redirect('orden_create')
        
    # GET: Cargar formulario
    servicios = Servicio.objects.all()
    servicios_data = [
        {
            'id': s.id,
            'nombre': s.nombre,
            'tarifa': float(s.tarifa),
            'tipo_cobro': s.tipo_cobro
        }
        for s in servicios
    ]
    
    promociones_activas = Promocion.objects.filter(
        esta_activa=True,
        fecha_inicio__lte=today,
        fecha_fin__ge=today
    )
    
    prendas_catalog = [p.nombre for p in CatalogoPrenda.objects.all().order_by('nombre')]
    
    context = {
        'clientes_frecuentes': Cliente.objects.annotate(num_ordenes=Count('ordenes')).order_by('-num_ordenes')[:5],
        'servicios_json': json.dumps(servicios_data),
        'active_tab': 'ordenes',
        'promociones': promociones_activas,
        'prendas_catalog_json': json.dumps(prendas_catalog),
    }
    return render(request, 'nueva_orden.html', context)


@login_required
@group_required('Administrador', 'Cajero')
def orden_create_partial(request, cliente_id):
    from .models import CatalogoPrenda, Promocion
    from django.utils import timezone
    today = timezone.localtime(timezone.now()).date()

    try:
        cliente = Cliente.objects.get(id=cliente_id)
    except Cliente.DoesNotExist:
        return HttpResponse("<script>showToast('El cliente seleccionado no existe.', 'error')</script>", status=404)

    if request.method == 'POST':
        observaciones = request.POST.get('observaciones', '')
        promocion_id = request.POST.get('promocion_id')
        
        promocion = None
        if promocion_id:
            try:
                promocion = Promocion.objects.filter(id=promocion_id).first()
            except Exception:
                pass

        try:
            total = Decimal(request.POST.get('total', '0.00') or '0.00')
            descuento = Decimal(request.POST.get('descuento', '0.00') or '0.00')
            total_a_pagar = Decimal(request.POST.get('total_a_pagar', '0.00') or '0.00')
            anticipo = Decimal(request.POST.get('anticipo', '0.00') or '0.00')
        except (InvalidOperation, TypeError, ValueError):
            return HttpResponse("<script>showToast('Error al procesar los montos de la orden.', 'error')</script>", status=400)
            
        metodo_pago = request.POST.get('metodo_pago', 'Efectivo')
        prendas_json = request.POST.get('prendas_json', '[]')
        fecha_entrega_estimada_str = request.POST.get('fecha_entrega_estimada')
        
        # Validar turno de caja antes de hacer cualquier cambio
        turno_activo = TurnoCaja.objects.filter(estado='Abierta').first()
        if not turno_activo:
            return HttpResponse("<script>showToast('No puedes registrar órdenes nuevas sin abrir un turno de caja primero.', 'error')</script>", status=400)
        
        try:
            prendas_data = json.loads(prendas_json)
        except json.JSONDecodeError:
            return HttpResponse("<script>showToast('Error en el formato de los detalles de las prendas.', 'error')</script>", status=400)
            
        if not prendas_data:
            return HttpResponse("<script>showToast('Debe agregar al menos una prenda a la orden.', 'error')</script>", status=400)
        
        fecha_entrega_estimada = None
        if fecha_entrega_estimada_str:
            fecha_entrega_estimada = parse_datetime(fecha_entrega_estimada_str)
            if fecha_entrega_estimada and timezone.is_naive(fecha_entrega_estimada):
                fecha_entrega_estimada = timezone.make_aware(fecha_entrega_estimada)
        
        try:
            with transaction.atomic():
                orden = Orden.objects.create(
                    cliente=cliente,
                    promocion=promocion,
                    fecha_entrega_estimada=fecha_entrega_estimada,
                    total=total,
                    descuento=descuento,
                    total_a_pagar=total_a_pagar,
                    anticipo=anticipo,
                    saldo_pendiente=total_a_pagar - anticipo,
                    estado_pago='Liquidado' if anticipo >= total_a_pagar else ('Parcial' if anticipo > 0 else 'Pendiente'),
                    estado_actual='Recibida',
                    observaciones=observaciones or None
                )
                
                SeguimientoLavado.objects.create(
                    orden=orden,
                    estado='Recibida',
                    observaciones='Ingreso inicial de la orden de servicio al sistema.'
                )
                
                for p in prendas_data:
                    servicio = Servicio.objects.get(id=p['servicio_id'])
                    tipo_prenda_nombre = p['tipo_prenda'].strip()
                    if tipo_prenda_nombre:
                        CatalogoPrenda.objects.get_or_create(nombre=tipo_prenda_nombre)
                    precio_unitario = float(p['precio_unitario']) if p.get('precio_unitario') is not None else None
                    PrendaOrden.objects.create(
                        orden=orden,
                        tipo_prenda=tipo_prenda_nombre,
                        cantidad=int(p.get('cantidad', 1)),
                        peso=float(p['peso']) if p.get('peso') else None,
                        servicio=servicio,
                        precio_unitario=precio_unitario,
                        es_delicada=bool(p.get('es_delicada', False)),
                        observaciones=p.get('observaciones', '') or None
                    )
                
                if anticipo > 0:
                    Pago.objects.create(
                        orden=orden,
                        monto=anticipo,
                        metodo_pago=metodo_pago,
                        tipo_pago='Anticipo',
                        observaciones='Abono inicial al registrar la orden',
                        turno=turno_activo
                    )
            
            # Recuperar historial actualizado del cliente
            ordenes = Orden.objects.filter(cliente=cliente).order_by('-fecha_recepcion').prefetch_related('prendas__servicio')
            
            context = {
                'cliente': cliente,
                'ordenes': ordenes,
                'success_message': "Orden de servicio registrada con éxito."
            }
            # Retornar el fragmento del historial para que se renderice en el panel derecho
            response = render(request, 'partials/cliente_historial.html', context)
            # Agregar el script para disparar la notificación Toast
            # Nota: El HTML retornado incluirá al final un script para activar el toast
            return response
            
        except Exception as e:
            return HttpResponse(f"<script>showToast('Error al procesar la orden: {str(e)}', 'error')</script>", status=500)

    # GET: Cargar formulario parcial
    servicios = Servicio.objects.all()
    servicios_data = [
        {
            'id': s.id,
            'nombre': s.nombre,
            'tarifa': float(s.tarifa),
            'tipo_cobro': s.tipo_cobro
        }
        for s in servicios
    ]
    
    promociones_activas = Promocion.objects.filter(
        esta_activa=True,
        fecha_inicio__lte=today,
        fecha_fin__gte=today
    )
    
    promociones_data = [
        {
            'id': p.id,
            'tipo': p.tipo,
            'valor': float(p.valor)
        }
        for p in promociones_activas
    ]
    
    prendas_catalog = [p.nombre for p in CatalogoPrenda.objects.all().order_by('nombre')]
    
    context = {
        'cliente': cliente,
        'servicios_json': json.dumps(servicios_data),
        'servicios': servicios,
        'promociones': promociones_activas,
        'promociones_json': json.dumps(promociones_data),
        'prendas_catalog_json': json.dumps(prendas_catalog),
    }
    return render(request, 'partials/orden_create_partial.html', context)


# ==========================================
#  DETALLE DE ORDEN Y SEGUIMIENTO (PASO 7)
# ==========================================

@login_required
@group_required('Administrador', 'Cajero')
def orden_detail(request, orden_id):
    try:
        orden = Orden.objects.select_related('cliente').get(id=orden_id)
    except Orden.DoesNotExist:
        messages.error(request, "La orden solicitada no existe.")
        return redirect('ordenes_list')

    prendas = PrendaOrden.objects.filter(orden=orden).select_related('servicio')
    seguimientos = SeguimientoLavado.objects.filter(orden=orden).order_by('fecha_cambio')
    pagos = Pago.objects.filter(orden=orden).order_by('fecha_pago')

    # Determinar el siguiente estado disponible
    estado_actual = orden.estado_actual
    siguiente_estado = None
    if estado_actual in ESTADOS_LAVADO:
        idx = ESTADOS_LAVADO.index(estado_actual)
        if idx < len(ESTADOS_LAVADO) - 1:
            siguiente_estado = ESTADOS_LAVADO[idx + 1]

    # Calcular subtotal por prenda
    prendas_detalle = []
    for p in prendas:
        tarifa = float(p.precio_unitario if p.precio_unitario is not None else p.servicio.tarifa)
        if 'kg' in p.servicio.tipo_cobro.lower():
            subtotal = tarifa * float(p.peso or 0)
        else:
            subtotal = tarifa * p.cantidad
        prendas_detalle.append({
            'prenda': p,
            'subtotal': subtotal
        })

    turno_activo = TurnoCaja.objects.filter(estado='Abierta').first()
    efectivo_disponible = turno_activo.efectivo_disponible if turno_activo else Decimal('0.00')

    context = {
        'orden': orden,
        'prendas_detalle': prendas_detalle,
        'seguimientos': seguimientos,
        'pagos': pagos,
        'siguiente_estado': siguiente_estado,
        'estados_lavado': ESTADOS_LAVADO,
        'estado_actual_idx': ESTADOS_LAVADO.index(estado_actual) if estado_actual in ESTADOS_LAVADO else 0,
        'efectivo_disponible': efectivo_disponible,
    }

    return render(request, 'orden_detalle.html', context)



@login_required
@group_required('Administrador', 'Cajero')
def orden_register_payment(request, orden_id):
    if request.method == 'POST':
        try:
            orden = Orden.objects.get(id=orden_id)
        except Orden.DoesNotExist:
            return HttpResponse("<p style='color:var(--danger);'>Orden no encontrada.</p>", status=404)

        # Validar si el turno de caja está abierto antes de registrar cobros
        turno_activo = TurnoCaja.objects.filter(estado='Abierta').first()
        if not turno_activo:
            pagos = Pago.objects.filter(orden=orden).order_by('fecha_pago')
            return render(request, 'partials/orden_payment_panel.html', {
                'orden': orden,
                'pagos': pagos,
                'payment_error': 'No se pueden registrar cobros si el turno de caja está cerrado. Abre un turno en el módulo de Caja primero.'
            })

        try:
            monto_str = request.POST.get('monto', '0.00').strip()
            monto = Decimal(monto_str or '0.00')
        except InvalidOperation:
            monto = Decimal('0.00')

        metodo_pago = request.POST.get('metodo_pago', 'Efectivo')
        observaciones_pago = request.POST.get('observaciones_pago', '').strip()
        confirmacion_cambio = request.POST.get('confirmar_cambio')

        pagos = Pago.objects.filter(orden=orden).order_by('fecha_pago')

        if monto <= 0:
            return render(request, 'partials/orden_payment_panel.html', {
                'orden': orden, 'pagos': pagos, 'payment_error': 'El monto debe ser mayor a $0.00.'
            })

        saldo_actual = Decimal(str(orden.saldo_pendiente))
        
        # Si la orden ya está liquidada (saldo 0) y llega un cobro redundante, lo ignoramos sin lanzar error
        if saldo_actual <= 0:
            return render(request, 'partials/orden_payment_panel.html', {'orden': orden, 'pagos': pagos})

        # Autocorrección: si por alguna razón técnica el monto supera el saldo real, simplemente lo limitamos
        if monto > saldo_actual:
            monto = saldo_actual
        
        monto_registrado = monto
        observaciones_db = observaciones_pago or f"Pago registrado (${monto_registrado:.2f})"

        with transaction.atomic():
            tipo_pago = 'Saldo Final' if monto_registrado >= saldo_actual else 'Abono'
            turno_activo = TurnoCaja.objects.filter(estado='Abierta').first()

            Pago.objects.create(
                orden=orden,
                monto=monto_registrado,
                metodo_pago=metodo_pago,
                tipo_pago=tipo_pago,
                observaciones=observaciones_db,
                turno=turno_activo
            )

            nuevo_anticipo = Decimal(str(orden.anticipo)) + monto_registrado
            nuevo_saldo = Decimal(str(orden.total_a_pagar)) - nuevo_anticipo

            orden.anticipo = nuevo_anticipo
            orden.saldo_pendiente = max(nuevo_saldo, Decimal('0.00'))

            if orden.saldo_pendiente <= Decimal('0.00'):
                orden.estado_pago = 'Liquidado'
            elif nuevo_anticipo > Decimal('0.00'):
                orden.estado_pago = 'Parcial'

            orden.save()

        pagos = Pago.objects.filter(orden=orden).order_by('fecha_pago')
        
        # Calcular efectivo disponible actualizado después de registrar el pago
        efectivo_disponible = turno_activo.efectivo_disponible if turno_activo else Decimal('0.00')

        context = {
            'orden': orden,
            'pagos': pagos,
            'efectivo_disponible': efectivo_disponible,
            'cambio_info': {
                'monto_cobrado': monto_registrado,
                'efectivo_recibido': efectivo_recibido,
                'cambio': cambio
            } if cambio > 0 else None
        }

        response = render(request, 'partials/orden_payment_panel.html', context)
        response['HX-Trigger'] = 'paymentUpdated'
        return response

    return HttpResponse(status=405)


@login_required
@group_required('Administrador', 'Cajero')
def pagos_list(request):
    local_now = timezone.localtime(timezone.now())
    start_of_today = local_now.replace(hour=0, minute=0, second=0, microsecond=0)
    end_of_today = local_now.replace(hour=23, minute=59, second=59, microsecond=999999)

    pagos = Pago.objects.select_related('orden__cliente').order_by('-fecha_pago')

    pagos_hoy = Pago.objects.filter(fecha_pago__range=(start_of_today, end_of_today))
    
    total_recaudado = pagos_hoy.aggregate(Sum('monto'))['monto__sum'] or Decimal('0.00')
    
    pagos_hoy_efectivo = pagos_hoy.filter(metodo_pago='Efectivo')
    total_efectivo = pagos_hoy_efectivo.aggregate(Sum('monto'))['monto__sum'] or Decimal('0.00')
    
    total_transferencia = pagos_hoy.filter(metodo_pago='Transferencia').aggregate(Sum('monto'))['monto__sum'] or Decimal('0.00')
    total_tarjeta = pagos_hoy.filter(metodo_pago='Tarjeta').aggregate(Sum('monto'))['monto__sum'] or Decimal('0.00')

    # Calcular cuánto efectivo salió de la caja como "cambio" hoy
    total_cambio_entregado = Decimal('0.00')
    for p in pagos_hoy_efectivo:
        if p.observaciones and 'Cambio entregado: $' in p.observaciones:
            try:
                m = re.search(r'Cambio entregado: \$([\d\.]+)', p.observaciones)
                if m:
                    total_cambio_entregado += Decimal(m.group(1))
            except (InvalidOperation, ValueError):
                pass

    context = {
        'pagos': pagos,
        'total_recaudado': total_recaudado,
        'total_efectivo': total_efectivo,
        'total_transferencia': total_transferencia,
        'total_tarjeta': total_tarjeta,
        'total_cambio_entregado': total_cambio_entregado,
    }
    return render(request, 'pagos.html', context)


@login_required
@group_required('Administrador', 'Cajero')
def pagos_search(request):
    query = request.GET.get('search', '').strip()
    metodo = request.GET.get('metodo_pago', '').strip()
    tipo = request.GET.get('tipo_pago', '').strip()
    fecha_rango = request.GET.get('fecha_rango', '').strip()

    pagos = Pago.objects.select_related('orden__cliente').all().order_by('-fecha_pago')

    if query:
        id_filter = Q()
        if query.isdigit():
            id_filter = Q(orden__id=int(query))
        pagos = pagos.filter(
            id_filter | Q(orden__cliente__nombre__icontains=query)
        )

    if metodo:
        pagos = pagos.filter(metodo_pago=metodo)

    if tipo:
        pagos = pagos.filter(tipo_pago=tipo)

    if fecha_rango:
        if " to " in fecha_rango:
            try:
                start_str, end_str = fecha_rango.split(" to ")
                start_date = timezone.datetime.strptime(start_str, "%Y-%m-%d")
                end_date = timezone.datetime.strptime(end_str, "%Y-%m-%d")
                start_datetime = timezone.make_aware(start_date.replace(hour=0, minute=0, second=0, microsecond=0))
                end_datetime = timezone.make_aware(end_date.replace(hour=23, minute=59, second=59, microsecond=999999))
                pagos = pagos.filter(fecha_pago__range=(start_datetime, end_datetime))
            except ValueError:
                pass
        else:
            try:
                date_obj = timezone.datetime.strptime(fecha_rango, "%Y-%m-%d")
                start_datetime = timezone.make_aware(date_obj.replace(hour=0, minute=0, second=0, microsecond=0))
                end_datetime = timezone.make_aware(date_obj.replace(hour=23, minute=59, second=59, microsecond=999999))
                pagos = pagos.filter(fecha_pago__range=(start_datetime, end_datetime))
            except ValueError:
                pass

    return render(request, 'partials/pagos_table.html', {'pagos': pagos})


@login_required
@group_required('Administrador', 'Cajero')
def orden_ticket(request, orden_id):
    try:
        orden = Orden.objects.select_related('cliente').get(id=orden_id)
    except Orden.DoesNotExist:
        messages.error(request, "La orden solicitada no existe.")
        return redirect('ordenes_list')

    prendas = PrendaOrden.objects.filter(orden=orden).select_related('servicio')
    pagos = Pago.objects.filter(orden=orden).order_by('fecha_pago')

    prendas_detalle = []
    for p in prendas:
        tarifa = float(p.precio_unitario if p.precio_unitario is not None else p.servicio.tarifa)
        if 'kg' in p.servicio.tipo_cobro.lower():
            subtotal = tarifa * float(p.peso or 0)
        else:
            subtotal = tarifa * p.cantidad
        prendas_detalle.append({
            'prenda': p,
            'subtotal': subtotal
        })

    context = {
        'orden': orden,
        'prendas_detalle': prendas_detalle,
        'pagos': pagos,
    }
    return render(request, 'orden_ticket.html', context)

# ==========================================
# GESTIÓN DE USUARIOS
# ==========================================

@login_required
@group_required('Administrador')
def usuarios_list(request):
    usuarios = User.objects.all().order_by('-is_active', 'first_name', 'username').prefetch_related('groups')
    total_usuarios = usuarios.count()
    activos_count = usuarios.filter(is_active=True).count()
    inactivos_count = usuarios.filter(is_active=False).count()
    context = {
        'usuarios': usuarios,
        'total_usuarios': total_usuarios,
        'activos_count': activos_count,
        'inactivos_count': inactivos_count,
        'active_tab': 'usuarios'
    }
    return render(request, 'usuarios.html', context)

@login_required
@group_required('Administrador')
def usuarios_search(request):
    q = request.GET.get('q', '').strip()
    estado = request.GET.get('estado', '')
    
    usuarios = User.objects.all().order_by('-is_active', 'first_name', 'username').prefetch_related('groups')
    
    if q:
        usuarios = usuarios.filter(Q(username__icontains=q) | Q(first_name__icontains=q) | Q(last_name__icontains=q) | Q(email__icontains=q))
    if estado == 'activos':
        usuarios = usuarios.filter(is_active=True)
    elif estado == 'inactivos':
        usuarios = usuarios.filter(is_active=False)
        
    return render(request, 'partials/usuarios_table.html', {'usuarios': usuarios})

@login_required
@group_required('Administrador')
def usuario_create(request):
    grupos = Group.objects.all()
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        confirm_password = request.POST.get('confirm_password')
        first_name = request.POST.get('first_name')
        last_name = request.POST.get('last_name')
        email = request.POST.get('email')
        rol_id = request.POST.get('rol')
        
        if password != confirm_password:
            messages.error(request, 'Las contraseñas no coinciden.')
        elif User.objects.filter(username=username).exists():
            messages.error(request, 'El nombre de usuario ya existe.')
        else:
            user = User.objects.create_user(
                username=username,
                password=password,
                email=email,
                first_name=first_name,
                last_name=last_name
            )
            if rol_id:
                grupo = Group.objects.get(id=rol_id)
                user.groups.add(grupo)
            messages.success(request, 'Usuario creado exitosamente.')
            return redirect('usuarios_list')
            
    context = {
        'grupos': grupos,
        'active_tab': 'usuarios'
    }
    return render(request, 'usuario_form.html', context)

@login_required
@group_required('Administrador')
def usuario_update(request, id):
    try:
        usuario = User.objects.get(id=id)
    except User.DoesNotExist:
        messages.error(request, 'El usuario solicitado no existe.')
        return redirect('usuarios_list')
    
    grupos = Group.objects.all()
    if request.method == 'POST':
        new_username = request.POST.get('username')
        if new_username != usuario.username and User.objects.filter(username=new_username).exists():
            messages.error(request, 'El nombre de usuario ya existe.')
            return redirect('usuario_update', id=id)

        usuario.username = new_username
        usuario.first_name = request.POST.get('first_name')
        usuario.last_name = request.POST.get('last_name')
        usuario.email = request.POST.get('email')
        
        password = request.POST.get('password')
        if password:
            confirm_password = request.POST.get('confirm_password')
            if password == confirm_password:
                usuario.set_password(password)
            else:
                messages.error(request, 'Las contraseñas no coinciden.')
                return redirect('usuario_update', id=id)
                
        rol_id = request.POST.get('rol')
        if rol_id:
            usuario.groups.clear()
            grupo = Group.objects.get(id=rol_id)
            usuario.groups.add(grupo)
            
        usuario.save()
        messages.success(request, 'Usuario actualizado exitosamente.')
        return redirect('usuarios_list')
        
    context = {
        'usuario': usuario,
        'grupos': grupos,
        'active_tab': 'usuarios'
    }
    return render(request, 'usuario_form.html', context)

@login_required
@group_required('Administrador')
def usuario_toggle_status(request, id):
    if request.method == 'POST':
        try:
            usuario = User.objects.get(id=id)
        except User.DoesNotExist:
            messages.error(request, 'El usuario solicitado no existe.')
            return redirect('usuarios_list')
        
        if usuario == request.user:
            messages.error(request, 'No puedes desactivar tu propia cuenta.')
            return redirect('usuarios_list')
            
        usuario.is_active = not usuario.is_active
        usuario.save()
        estado = "activado" if usuario.is_active else "desactivado"
        messages.success(request, f'Usuario {estado} exitosamente.')
    return redirect('usuarios_list')

# ==========================================
# SEGUIMIENTO (KANBAN)
# ==========================================
@login_required
@group_required('Administrador', 'Cajero', 'Operador')
def seguimiento_list(request):
    ordenes_activas = Orden.objects.exclude(estado_actual='Entregada').select_related('cliente').prefetch_related('prendas').order_by('fecha_recepcion')
    
    # Parámetro para seguimiento individual o tomar la primera activa por defecto
    orden_id = request.GET.get('orden_id')
    orden_seleccionada = None
    seguimientos_orden = []
    
    if orden_id:
        try:
            orden_seleccionada = Orden.objects.select_related('cliente').prefetch_related('prendas').get(id=orden_id)
        except Orden.DoesNotExist:
            pass
            
    # Si no se seleccionó ninguna en particular pero hay activas, precargar la primera
    if not orden_seleccionada and ordenes_activas.exists():
        orden_seleccionada = ordenes_activas.first()
        
    if orden_seleccionada:
        seguimientos_orden = SeguimientoLavado.objects.filter(orden=orden_seleccionada).order_by('fecha_cambio')
        
    context = {
        'ordenes_activas': ordenes_activas,
        'orden_seleccionada': orden_seleccionada,
        'seguimientos_orden': seguimientos_orden,
    }
    return render(request, 'seguimiento.html', context)

@login_required
@group_required('Administrador', 'Cajero', 'Operador')
def seguimiento_advance_status(request, orden_id):
    if request.method == 'POST':
        from django.db import transaction
        from django.contrib import messages
        from django.shortcuts import redirect
        
        orden = get_object_or_404(Orden, id=orden_id)
        estados = ['Recibida', 'Clasificada', 'En lavado', 'En secado', 'En planchado', 'Empaquetada', 'Lista para entrega', 'Entregada']
        try:
            idx = estados.index(orden.estado_actual)
            if idx < len(estados) - 1:
                nuevo_estado = estados[idx + 1]
                
                # Validación contable: No entregar si tiene saldo pendiente
                if nuevo_estado == 'Entregada' and orden.saldo_pendiente > 0:
                    messages.error(request, f"No se puede entregar la orden #{orden.id:06d} porque tiene un saldo pendiente de ${orden.saldo_pendiente:.2f}. El pago total es obligatorio.")
                else:
                    with transaction.atomic():
                        orden.estado_actual = nuevo_estado
                        if nuevo_estado == 'Entregada':
                            from django.utils import timezone
                            orden.fecha_entrega_real = timezone.now()
                        orden.save()
                        
                        # Registrar en el historial de seguimiento
                        SeguimientoLavado.objects.create(
                            orden=orden,
                            estado=nuevo_estado,
                            observaciones=f"Estado avanzado a {nuevo_estado} desde la vista de seguimiento."
                        )
                        messages.success(request, f"Orden #{orden.id:06d} avanzada a {nuevo_estado} con éxito.")
        except ValueError:
            pass
            
        from django.urls import reverse
        return redirect(reverse('seguimiento_list') + f'?orden_id={orden.id}')
    
    return HttpResponse(status=405)


# ==========================================
#  MÓDULO DE REPORTES
# ==========================================

def get_report_data(periodo, start_date=None, end_date=None):
    local_now = timezone.localtime(timezone.now())
    
    fecha_inicio = None
    fecha_fin = None
    titulo_periodo = ""
    
    if start_date and end_date:
        try:
            from django.utils.dateparse import parse_date
            parsed_start = parse_date(start_date)
            parsed_end = parse_date(end_date)
            if parsed_start and parsed_end:
                fecha_inicio = timezone.make_aware(timezone.datetime.combine(parsed_start, timezone.datetime.min.time()))
                fecha_fin = timezone.make_aware(timezone.datetime.combine(parsed_end, timezone.datetime.max.time()))
                titulo_periodo = f"Personalizado: {parsed_start.strftime('%d/%m/%Y')} al {parsed_end.strftime('%d/%m/%Y')}"
                periodo = 'rango'
        except Exception:
            pass

    if not titulo_periodo:
        if periodo == 'hoy':
            fecha_inicio = local_now.replace(hour=0, minute=0, second=0, microsecond=0)
            fecha_fin = local_now.replace(hour=23, minute=59, second=59, microsecond=999999)
            titulo_periodo = "Hoy"
        elif periodo == 'semana':
            dias_restar = local_now.weekday()
            fecha_inicio = (local_now - timezone.timedelta(days=dias_restar)).replace(hour=0, minute=0, second=0, microsecond=0)
            fecha_fin = local_now.replace(hour=23, minute=59, second=59, microsecond=999999)
            titulo_periodo = "Esta Semana"
        elif periodo == 'historico':
            fecha_inicio = None
            fecha_fin = None
            titulo_periodo = "Histórico (Todo)"
        else: # mes (default)
            fecha_inicio = local_now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
            fecha_fin = local_now.replace(hour=23, minute=59, second=59, microsecond=999999)
            titulo_periodo = "Este Mes"
            periodo = 'mes'
        
    pagos_filter = Q()
    ordenes_filter = Q()
    if fecha_inicio and fecha_fin:
        pagos_filter = Q(fecha_pago__range=(fecha_inicio, fecha_fin))
        ordenes_filter = Q(fecha_recepcion__range=(fecha_inicio, fecha_fin))

    pagos = Pago.objects.filter(pagos_filter)
    ingresos_totales = pagos.aggregate(Sum('monto'))['monto__sum'] or 0.00
    
    ingresos_efectivo = pagos.filter(metodo_pago='Efectivo').aggregate(Sum('monto'))['monto__sum'] or 0.00
    ingresos_transferencia = pagos.filter(metodo_pago='Transferencia').aggregate(Sum('monto'))['monto__sum'] or 0.00
    ingresos_tarjeta = pagos.filter(metodo_pago='Tarjeta').aggregate(Sum('monto'))['monto__sum'] or 0.00

    ordenes_por_cobrar = Orden.objects.filter(saldo_pendiente__gt=0).exclude(estado_actual='Entregada')
    total_por_cobrar = ordenes_por_cobrar.aggregate(Sum('saldo_pendiente'))['saldo_pendiente__sum'] or 0.00
    top_deudores = ordenes_por_cobrar.order_by('-saldo_pendiente')[:5]

    ordenes = Orden.objects.filter(ordenes_filter)
    total_ordenes = ordenes.count()
    
    ordenes_por_estado = ordenes.values('estado_actual').annotate(total=Count('id')).order_by('-total')

    if fecha_inicio and fecha_fin:
        top_clientes = Cliente.objects.annotate(
            total_gastado=Sum('ordenes__total_a_pagar', filter=Q(ordenes__fecha_recepcion__range=(fecha_inicio, fecha_fin))),
            total_ordenes=Count('ordenes', filter=Q(ordenes__fecha_recepcion__range=(fecha_inicio, fecha_fin)))
        ).exclude(total_gastado=None).order_by('-total_gastado')[:5]
    else:
        top_clientes = Cliente.objects.annotate(
            total_gastado=Sum('ordenes__total_a_pagar'),
            total_ordenes=Count('ordenes')
        ).exclude(total_gastado=None).order_by('-total_gastado')[:5]

    servicios_stats = PrendaOrden.objects.filter(
        orden__fecha_recepcion__range=(fecha_inicio, fecha_fin) if fecha_inicio else Q()
    ).values('servicio__nombre').annotate(
        cantidad_solicitada=Sum('cantidad')
    ).order_by('-cantidad_solicitada')[:5]

    return {
        'periodo': periodo,
        'titulo_periodo': titulo_periodo,
        'ingresos_totales': ingresos_totales,
        'ingresos_efectivo': ingresos_efectivo,
        'ingresos_transferencia': ingresos_transferencia,
        'ingresos_tarjeta': ingresos_tarjeta,
        'total_por_cobrar': total_por_cobrar,
        'top_deudores': top_deudores,
        'total_ordenes': total_ordenes,
        'ordenes_por_estado': ordenes_por_estado,
        'top_clientes': top_clientes,
        'servicios_stats': servicios_stats,
        'fecha_inicio_str': start_date or '',
        'fecha_fin_str': end_date or '',
    }

@login_required
@group_required('Administrador')
def reportes_dashboard(request):
    periodo = request.GET.get('periodo', 'mes')
    fecha_inicio = request.GET.get('fecha_inicio', '').strip()
    fecha_fin = request.GET.get('fecha_fin', '').strip()
    context = get_report_data(periodo, fecha_inicio, fecha_fin)
    return render(request, 'reportes.html', context)

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
import io
try:
    from xhtml2pdf import pisa
except ImportError:
    pisa = None
from django.template.loader import get_template

@login_required
@group_required('Administrador')
def reportes_export_excel(request):
    periodo = request.GET.get('periodo', 'mes')
    fecha_inicio = request.GET.get('fecha_inicio', '').strip()
    fecha_fin = request.GET.get('fecha_fin', '').strip()
    data = get_report_data(periodo, fecha_inicio, fecha_fin)
    
    wb = openpyxl.Workbook()
    
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2563EB")
    header_align = Alignment(horizontal="center", vertical="center")
    
    ws = wb.active
    ws.title = "Resumen Financiero"
    
    ws.append([f"Reporte Financiero - {data['titulo_periodo']}"])
    ws.append([])
    ws.append(["Métrica", "Valor"])
    
    for row in ws.iter_rows(min_row=3, max_row=3):
        for cell in row:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            
    ws.append(["Ingresos Totales", f"${data['ingresos_totales']:.2f}"])
    ws.append(["Efectivo", f"${data['ingresos_efectivo']:.2f}"])
    ws.append(["Transferencia", f"${data['ingresos_transferencia']:.2f}"])
    ws.append(["Tarjeta", f"${data['ingresos_tarjeta']:.2f}"])
    ws.append(["Cuentas por Cobrar", f"${data['total_por_cobrar']:.2f}"])
    ws.append(["Total de Órdenes", data['total_ordenes']])
    
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 15

    ws2 = wb.create_sheet(title="Top Clientes")
    ws2.append(["Cliente", "Órdenes", "Total Gastado"])
    for row in ws2.iter_rows(min_row=1, max_row=1):
        for cell in row:
            cell.font = header_font
            cell.fill = header_fill
    for c in data['top_clientes']:
        ws2.append([c.nombre_completo, c.total_ordenes, f"${c.total_gastado:.2f}"])
    ws2.column_dimensions['A'].width = 30
    ws2.column_dimensions['C'].width = 15

    ws3 = wb.create_sheet(title="Cuentas por Cobrar")
    ws3.append(["Orden", "Cliente", "Teléfono", "Saldo Pendiente"])
    for row in ws3.iter_rows(min_row=1, max_row=1):
        for cell in row:
            cell.font = header_font
            cell.fill = header_fill
    for d in data['top_deudores']:
        ws3.append([f"#{d.id:06d}", d.cliente.nombre_completo, d.cliente.telefono, f"${d.saldo_pendiente:.2f}"])
    ws3.column_dimensions['A'].width = 15
    ws3.column_dimensions['B'].width = 30
    ws3.column_dimensions['C'].width = 15
    ws3.column_dimensions['D'].width = 20
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="reporte_lavafacil_{periodo}.xlsx"'
    wb.save(response)
    return response

@login_required
@group_required('Administrador')
def reportes_export_pdf(request):
    periodo = request.GET.get('periodo', 'mes')
    fecha_inicio = request.GET.get('fecha_inicio', '').strip()
    fecha_fin = request.GET.get('fecha_fin', '').strip()
    data = get_report_data(periodo, fecha_inicio, fecha_fin)
    
    from .models import Configuracion
    data['config'] = Configuracion.load()
    
    # Pre-calcular listas JSON para que Chart.js las lea fácilmente en el HTML
    data['chart_labels_dona'] = json.dumps([s['servicio__nombre'] for s in data.get('servicios_stats', [])[:5]])
    data['chart_data_dona'] = json.dumps([s['cantidad_solicitada'] for s in data.get('servicios_stats', [])[:5]])
    
    data['chart_labels_barras'] = json.dumps([e['estado_actual'] for e in data.get('ordenes_por_estado', [])])
    data['chart_data_barras'] = json.dumps([e['total'] for e in data.get('ordenes_por_estado', [])])
    
    return render(request, 'reporte_pdf.html', data)


# ==========================================
# CONFIGURACIÓN DEL SISTEMA
# ==========================================
from .models import Configuracion
from .forms import ConfiguracionForm

@login_required
@group_required('Administrador')
def configuracion_update(request):
    config = Configuracion.load()
    
    if request.method == 'POST':
        form = ConfiguracionForm(request.POST, request.FILES, instance=config)
        if form.is_valid():
            form.save()
            messages.success(request, "Configuración del sistema actualizada correctamente.")
            return redirect('configuracion')
        else:
            messages.error(request, "Por favor, corrija los errores en el formulario.")
    else:
        form = ConfiguracionForm(instance=config)
        
    context = {
        'form': form,
        'active_tab': 'configuracion'
    }
    return render(request, 'configuracion.html', context)

# ==========================================
# MÓDULO DE CAJA (CORTE DE CAJA)
# ==========================================
from .models import TurnoCaja, MovimientoCaja

@login_required
@group_required('Administrador', 'Cajero')
def caja_dashboard(request):
    # Buscar si hay un turno abierto para el usuario actual o en general
    # Si queremos que la caja sea por sucursal (general), buscamos la última abierta
    turno_abierto = TurnoCaja.objects.filter(estado='Abierta').first()
    
    movimientos = []
    pagos = []
    
    if turno_abierto:
        movimientos = MovimientoCaja.objects.filter(turno=turno_abierto).order_by('-fecha')
        pagos = Pago.objects.filter(turno=turno_abierto).order_by('-fecha_pago')
        
        ingresos_ordenes = pagos.aggregate(Sum('monto'))['monto__sum'] or Decimal('0.00')
        ingresos_extra = movimientos.filter(tipo='Ingreso').aggregate(Sum('monto'))['monto__sum'] or Decimal('0.00')
        egresos = movimientos.filter(tipo='Egreso').aggregate(Sum('monto'))['monto__sum'] or Decimal('0.00')
        
        saldo_calculado = turno_abierto.saldo_inicial + ingresos_ordenes + ingresos_extra - egresos
    else:
        saldo_calculado = Decimal('0.00')
        ingresos_ordenes = Decimal('0.00')
        ingresos_extra = Decimal('0.00')
        egresos = Decimal('0.00')

    context = {
        'turno_abierto': turno_abierto,
        'movimientos': movimientos,
        'pagos': pagos,
        'ingresos_ordenes': ingresos_ordenes,
        'ingresos_extra': ingresos_extra,
        'egresos': egresos,
        'saldo_calculado': saldo_calculado,
        'active_tab': 'caja'
    }
    return render(request, 'caja.html', context)

@login_required
@group_required('Administrador', 'Cajero')
def caja_abrir(request):
    if request.method == 'POST':
        if TurnoCaja.objects.filter(estado='Abierta').exists():
            messages.error(request, 'Ya existe un turno de caja abierto.')
            return redirect('caja_dashboard')
            
        try:
            saldo_inicial = Decimal(request.POST.get('saldo_inicial', '0.00'))
        except InvalidOperation:
            saldo_inicial = Decimal('0.00')
            
        if saldo_inicial < Decimal('10.00'):
            messages.error(request, 'El turno de caja debe iniciarse con un fondo mínimo obligatorio de $10.00 para asegurar el cambio inicial.')
            return redirect('caja_dashboard')
            
        TurnoCaja.objects.create(
            usuario=request.user,
            saldo_inicial=saldo_inicial,
            estado='Abierta'
        )
        messages.success(request, f'Turno de caja abierto exitosamente con ${saldo_inicial}.')
        
    return redirect('caja_dashboard')

@login_required
@group_required('Administrador', 'Cajero')
def caja_cerrar(request):
    turno = TurnoCaja.objects.filter(estado='Abierta').first()
    if not turno:
        messages.error(request, 'No hay turno abierto para cerrar.')
        return redirect('caja_dashboard')
        
    if request.method == 'POST':
        try:
            saldo_real = Decimal(request.POST.get('saldo_final_real', '0.00'))
        except InvalidOperation:
            saldo_real = Decimal('0.00')
            
        if saldo_real < 0:
            messages.error(request, 'El saldo físico real en caja no puede ser negativo.')
            return redirect('caja_dashboard')
            
        observaciones = request.POST.get('observaciones', '')
        
        # Calcular saldo esperado
        pagos = Pago.objects.filter(turno=turno)
        ingresos_ordenes = pagos.aggregate(Sum('monto'))['monto__sum'] or Decimal('0.00')
        
        movimientos = MovimientoCaja.objects.filter(turno=turno)
        ingresos_extra = movimientos.filter(tipo='Ingreso').aggregate(Sum('monto'))['monto__sum'] or Decimal('0.00')
        egresos = movimientos.filter(tipo='Egreso').aggregate(Sum('monto'))['monto__sum'] or Decimal('0.00')
        
        saldo_esperado = turno.saldo_inicial + ingresos_ordenes + ingresos_extra - egresos
        
        turno.saldo_final_esperado = saldo_esperado
        turno.saldo_final_real = saldo_real
        turno.fecha_cierre = timezone.now()
        turno.estado = 'Cerrada'
        turno.observaciones = observaciones
        turno.save()
        
        messages.success(request, 'Turno de caja cerrado exitosamente.')
        
    return redirect('caja_dashboard')

@login_required
@group_required('Administrador', 'Cajero')
def caja_movimiento_crear(request):
    turno = TurnoCaja.objects.filter(estado='Abierta').first()
    if not turno:
        messages.error(request, 'Debe abrir un turno de caja antes de registrar movimientos.')
        return redirect('caja_dashboard')
        
    if request.method == 'POST':
        tipo = request.POST.get('tipo', 'Egreso')
        concepto = request.POST.get('concepto', '')
        try:
            monto = Decimal(request.POST.get('monto', '0.00'))
        except InvalidOperation:
            monto = Decimal('0.00')
            
        if monto > 0 and concepto:
            if tipo == 'Egreso':
                # Calcular saldo actual
                pagos = Pago.objects.filter(turno=turno)
                ingresos_ordenes = pagos.aggregate(Sum('monto'))['monto__sum'] or Decimal('0.00')
                
                movimientos = MovimientoCaja.objects.filter(turno=turno)
                ingresos_extra = movimientos.filter(tipo='Ingreso').aggregate(Sum('monto'))['monto__sum'] or Decimal('0.00')
                egresos = movimientos.filter(tipo='Egreso').aggregate(Sum('monto'))['monto__sum'] or Decimal('0.00')
                
                saldo_calculado = turno.saldo_inicial + ingresos_ordenes + ingresos_extra - egresos
                
                if monto > saldo_calculado:
                    messages.error(request, f'No se puede registrar el egreso. Fondos insuficientes en caja (Disponible: ${saldo_calculado:.2f}).')
                    return redirect('caja_dashboard')

            MovimientoCaja.objects.create(
                turno=turno,
                tipo=tipo,
                concepto=concepto,
                monto=monto
            )
            messages.success(request, f'{tipo} registrado correctamente.')
        else:
            messages.error(request, 'Por favor, ingrese un monto válido y un concepto.')
            
    return redirect('caja_dashboard')

# ==========================================
#  EXPORTACIONES A CSV
# ==========================================
import csv

@login_required
@group_required('Administrador', 'Cajero')
def export_clientes_csv(request):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="clientes.csv"'
    
    writer = csv.writer(response)
    writer.writerow(['ID', 'Cédula', 'Nombres', 'Apellidos', 'Teléfono', 'Correo', 'Dirección'])
    
    clientes = Cliente.objects.all().order_by('id')
    for c in clientes:
        writer.writerow([c.id, c.cedula, c.nombre, c.apellido, c.telefono, c.correo or '', c.direccion or ''])
        
    return response

@login_required
@group_required('Administrador', 'Cajero')
def export_ordenes_csv(request):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="ordenes.csv"'
    
    writer = csv.writer(response)
    writer.writerow(['ID', 'Cliente', 'Fecha Recepcion', 'Estado Actual', 'Estado Pago', 'Total', 'Anticipo', 'Saldo Pendiente'])
    
    ordenes = Orden.objects.select_related('cliente').all().order_by('-fecha_recepcion')
    for o in ordenes:
        writer.writerow([
            o.id, 
            o.cliente.nombre_completo, 
            timezone.localtime(o.fecha_recepcion).strftime("%Y-%m-%d %H:%M"), 
            o.estado_actual, 
            o.estado_pago, 
            o.total_a_pagar, 
            o.anticipo, 
            o.saldo_pendiente
        ])
        
    return response

@login_required
@group_required('Administrador', 'Cajero')
def export_pagos_csv(request):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="pagos.csv"'
    
    writer = csv.writer(response)
    writer.writerow(['ID Pago', 'Orden ID', 'Fecha', 'Metodo', 'Tipo', 'Monto', 'Cajero/Turno'])
    
    pagos = Pago.objects.select_related('orden', 'turno__usuario').all().order_by('-fecha_pago')
    for p in pagos:
        cajero = p.turno.usuario.username if p.turno and p.turno.usuario else ''
        writer.writerow([
            p.id, 
            p.orden.id, 
            timezone.localtime(p.fecha_pago).strftime("%Y-%m-%d %H:%M"), 
            p.metodo_pago, 
            p.tipo_pago, 
            p.monto, 
            cajero
        ])
        
    return response

# ==========================================
#  PERFIL DE USUARIO
# ==========================================
from django.contrib.auth.forms import PasswordChangeForm
from django.contrib.auth import update_session_auth_hash

@login_required
def mi_perfil(request):
    user = request.user
    if request.method == 'POST':
        if 'update_profile' in request.POST:
            user.first_name = request.POST.get('first_name', user.first_name)
            user.last_name = request.POST.get('last_name', user.last_name)
            user.email = request.POST.get('email', user.email)
            user.save()
            messages.success(request, 'Perfil actualizado correctamente.')
            return redirect('mi_perfil')
            
        elif 'update_password' in request.POST:
            form = PasswordChangeForm(user, request.POST)
            if form.is_valid():
                user = form.save()
                update_session_auth_hash(request, user)
                messages.success(request, 'Contraseña actualizada correctamente.')
                return redirect('mi_perfil')
            else:
                for field, errors in form.errors.items():
                    for error in errors:
                        messages.error(request, error)
    else:
        form = PasswordChangeForm(user)
        
    context = {
        'password_form': form
    }
    return render(request, 'mi_perfil.html', context)

# ==========================================
#  PROMOCIONES
# ==========================================
from datetime import datetime

@login_required
@group_required('Administrador', 'Cajero')
def promociones_list(request):
    promociones = Promocion.objects.all().order_by('-fecha_inicio')
    context = {'promociones': promociones}
    return render(request, 'promociones.html', context)

@login_required
@group_required('Administrador', 'Cajero')
def promociones_search(request):
    query = request.GET.get('search', '')
    if query:
        promociones = Promocion.objects.filter(nombre__icontains=query).order_by('-fecha_inicio')
    else:
        promociones = Promocion.objects.all().order_by('-fecha_inicio')
    return render(request, 'partials/promociones_table.html', {'promociones': promociones})

@login_required
@group_required('Administrador', 'Cajero')
def promocion_create(request):
    if request.method == 'POST':
        nombre = request.POST.get('nombre')
        tipo = request.POST.get('tipo')
        valor = request.POST.get('valor')
        fecha_inicio_str = request.POST.get('fecha_inicio')
        fecha_fin_str = request.POST.get('fecha_fin')
        
        try:
            fecha_inicio = datetime.strptime(fecha_inicio_str, '%Y-%m-%d').date()
            fecha_fin = datetime.strptime(fecha_fin_str, '%Y-%m-%d').date()
            
            Promocion.objects.create(
                nombre=nombre,
                tipo=tipo,
                valor=valor,
                fecha_inicio=fecha_inicio,
                fecha_fin=fecha_fin,
                esta_activa=True
            )
            messages.success(request, 'Promoción creada exitosamente.')
        except Exception as e:
            messages.error(request, f'Error al crear la promoción: {str(e)}')
            
    return redirect('promociones_list')

@login_required
@group_required('Administrador', 'Cajero')
def promocion_edit(request, id):
    promocion = get_object_or_404(Promocion, id=id)
    if request.method == 'POST':
        promocion.nombre = request.POST.get('nombre')
        promocion.tipo = request.POST.get('tipo')
        promocion.valor = request.POST.get('valor')
        
        try:
            promocion.fecha_inicio = datetime.strptime(request.POST.get('fecha_inicio'), '%Y-%m-%d').date()
            promocion.fecha_fin = datetime.strptime(request.POST.get('fecha_fin'), '%Y-%m-%d').date()
            promocion.save()
            messages.success(request, 'Promoción actualizada exitosamente.')
        except Exception as e:
            messages.error(request, f'Error al actualizar la promoción: {str(e)}')
            
    return redirect('promociones_list')

@login_required
@group_required('Administrador', 'Cajero')
def promocion_delete(request, id):
    promocion = get_object_or_404(Promocion, id=id)
    promocion.delete()
    messages.success(request, 'Promoción eliminada exitosamente.')
    return redirect('promociones_list')

@login_required
@group_required('Administrador', 'Cajero')
def promocion_toggle(request, id):
    promocion = get_object_or_404(Promocion, id=id)
    promocion.esta_activa = not promocion.esta_activa
    promocion.save()
    
    promociones = Promocion.objects.all().order_by('-fecha_inicio')
    response = render(request, 'partials/promociones_table.html', {'promociones': promociones})
    
    estado = "activada" if promocion.esta_activa else "desactivada"
    mensaje = f"Promoción '{promocion.nombre}' {estado} correctamente."
    response['HX-Trigger'] = json.dumps({'showToast': {'message': mensaje, 'type': 'success'}})
    
    return response

